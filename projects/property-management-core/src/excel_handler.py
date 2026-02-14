# -*- coding: utf-8 -*
"""
Excel 파일 읽기/쓰기를 위한 중앙 핸들러 모듈.

이 모듈은 pandas를 사용하여 Excel 파일과의 모든 상호작용을 관리합니다.
- 지정된 시트에서 데이터 읽기
- 데이터프레임을 지정된 시트에 쓰기
- 파일 및 시트 존재 여부 확인
- 스타일링 및 열 너비 조정 (필요시)

기존의 Google Sheets 연동 로직을 대체하며, 로컬 Excel 파일을
데이터 소스로 사용하도록 시스템을 마이그레이션하는 핵심 요소입니다.
"""
import os
import pandas as pd
from typing import List, Dict, Optional, Any
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

class ExcelHandler:
    """Excel 파일과의 상호작용을 관리하는 클래스."""

    def __init__(self, file_path: str):
        """
        ExcelHandler를 초기화합니다.

        Args:
            file_path (str): 처리할 Excel 파일의 경로.
        """
        self.file_path = file_path

    def sheet_exists(self, sheet_name: str) -> bool:
        """
        파일에 특정 시트가 존재하는지 확인합니다.

        Args:
            sheet_name (str): 확인할 시트의 이름.

        Returns:
            bool: 시트가 존재하면 True, 그렇지 않으면 False.
        """
        if not os.path.exists(self.file_path):
            return False
        try:
            workbook = load_workbook(self.file_path, read_only=True)
            return sheet_name in workbook.sheetnames
        except Exception:
            # 파일이 손상되었거나 엑셀 파일이 아닌 경우
            return False

    def read_sheet(self, sheet_name: str, header_row: int = 0) -> Optional[pd.DataFrame]:
        """
        특정 시트에서 데이터를 읽어 데이터프레임으로 반환합니다.

        Args:
            sheet_name (str): 읽을 시트의 이름.
            header_row (int): 헤더로 사용할 행 번호 (0-based).

        Returns:
            Optional[pd.DataFrame]: 시트 데이터를 담은 데이터프레임. 시트가 없으면 None.
        """
        if not self.sheet_exists(sheet_name):
            return None
        
        try:
            return pd.read_excel(self.file_path, sheet_name=sheet_name, header=header_row)
        except Exception as e:
            print(f"'{sheet_name}' 시트를 읽는 중 오류 발생: {e}")
            return None

    def write_data(self, sheet_name: str, data: pd.DataFrame, start_row: int = 0, start_col: int = 0, index: bool = False, header: bool = True):
        """
        데이터프레임을 특정 시트에 씁니다. 파일이나 시트가 없으면 새로 만듭니다.

        Args:
            sheet_name (str): 데이터를 쓸 시트의 이름.
            data (pd.DataFrame): 시트에 쓸 데이터.
            start_row (int): 쓰기를 시작할 행 위치 (0-based).
            start_col (int): 쓰기를 시작할 열 위치 (0-based).
            index (bool): 데이터프레임의 인덱스를 쓸지 여부.
            header (bool): 데이터프레임의 헤더를 쓸지 여부.
        """
        if not os.path.exists(self.file_path):
            # 파일이 없으면 새로 만들기
            with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
                data.to_excel(writer, sheet_name=sheet_name, startrow=start_row, startcol=start_col, index=index, header=header)
            self._auto_resize_columns(sheet_name)
            return

        try:
            # 기존 파일에 시트 추가 또는 덮어쓰기
            with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                # if_sheet_exists='replace'는 openpyxl 1.0.0, pandas 1.4.0 이상에서 지원
                # 기존 시트를 완전히 대체함
                data.to_excel(writer, sheet_name=sheet_name, startrow=start_row, startcol=start_col, index=index, header=header)

        except ValueError as e:
             # openpyxl < 1.0.0 에서는 if_sheet_exists 인자 지원 안함, 수동으로 시트 제거 후 다시 쓰기
             if "already exists" in str(e):
                book = load_workbook(self.file_path)
                if sheet_name in book.sheetnames:
                    del book[sheet_name]
                book.save(self.file_path)
                book.close()

                with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a') as writer:
                    data.to_excel(writer, sheet_name=sheet_name, startrow=start_row, startcol=start_col, index=index, header=header)
             else:
                raise e
        
        self._auto_resize_columns(sheet_name)

    def _auto_resize_columns(self, sheet_name: str):
        """
        지정된 시트의 모든 열 너비를 내용에 맞게 자동 조정합니다.
        
        Args:
            sheet_name (str): 열 너비를 조정할 시트의 이름.
        """
        try:
            book = load_workbook(self.file_path)
            sheet = book[sheet_name]
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width
            book.save(self.file_path)
        except Exception as e:
            print(f"'{sheet_name}' 시트의 열 너비 조정 중 오류 발생: {e}")

    def get_all_sheet_names(self) -> Optional[List[str]]:
        """
        Excel 파일의 모든 시트 이름을 리스트로 반환합니다.

        Returns:
            Optional[List[str]]: 시트 이름 리스트. 파일이 없으면 None.
        """
        if not os.path.exists(self.file_path):
            return None
        try:
            workbook = load_workbook(self.file_path, read_only=True)
            return workbook.sheetnames
        except Exception as e:
            print(f"파일에서 시트 목록을 가져오는 중 오류 발생: {e}")
            return None

if __name__ == '__main__':
    # 테스트용 코드
    TEST_FILE_PATH = 'test_excel_handler.xlsx'
    
    # 1. 핸들러 생성
    handler = ExcelHandler(TEST_FILE_PATH)
    
    # 2. 테스트 데이터 생성
    data1 = pd.DataFrame({
        '이름': ['홍길동', '이순신'],
        '나이': [30, 45],
        '도시': ['서울', '부산']
    })
    
    data2 = pd.DataFrame({
        '제품명': ['노트북', '마우스', '키보드'],
        '가격': [1500000, 50000, 75000],
        '재고': [10, 100, 50]
    })
    
    # 3. 데이터 쓰기 (새 파일, 새 시트)
    print("--- '고객정보' 시트에 데이터 쓰기 ---")
    handler.write_data('고객정보', data1)
    
    # 4. 다른 시트에 데이터 쓰기 (기존 파일)
    print("--- '제품재고' 시트에 데이터 추가 ---")
    handler.write_data('제품재고', data2)
    
    # 5. 시트 목록 확인
    print("\n--- 전체 시트 목록 ---")
    sheet_names = handler.get_all_sheet_names()
    print(f"시트: {sheet_names}")
    
    # 6. 특정 시트 존재 여부 확인
    print("\n--- 시트 존재 여부 확인 ---")
    print(f"'고객정보' 시트 존재 여부: {handler.sheet_exists('고객정보')}")
    print(f"'임시정보' 시트 존재 여부: {handler.sheet_exists('임시정보')}")
    
    # 7. 데이터 읽기
    print("\n--- '고객정보' 시트 데이터 읽기 ---")
    customer_df = handler.read_sheet('고객정보')
    if customer_df is not None:
        print(customer_df)
        
    # 8. 데이터 덮어쓰기
    print("\n--- '고객정보' 시트 덮어쓰기 ---")
    data1_updated = pd.DataFrame({
        'ID': ['C001', 'C002', 'C003'],
        '고객명': ['김철수', '박영희', '이지은'],
        '등급': ['Gold', 'Silver', 'Bronze']
    })
    handler.write_data('고객정보', data1_updated)
    
    # 9. 덮어쓴 데이터 확인
    print("\n--- 덮어쓴 '고객정보' 시트 데이터 다시 읽기 ---")
    customer_df_updated = handler.read_sheet('고객정보')
    if customer_df_updated is not None:
        print(customer_df_updated)

    # 10. 테스트 파일 삭제
    if os.path.exists(TEST_FILE_PATH):
        os.remove(TEST_FILE_PATH)
        print(f"\n테스트 파일 '{TEST_FILE_PATH}' 삭제 완료.")
